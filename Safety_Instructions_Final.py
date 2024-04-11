import os  
import tkinter as tk  
from tkinter import filedialog
from tkinter import BOTH
from tkinter import ttk  
from tkinter.ttk import Frame, Label, Style
from itertools import cycle  
from PIL import Image, ImageTk  
from openpyxl import load_workbook  
import time  
from openpyxl_image_loader import SheetImageLoader  

class SlideshowApp:
    def __init__(self, master, image_folder, text_file=None, sleep_time=3):
        # Constructor method for SlideshowApp class
        self.master = master  # Store the master window
        self.image_folder = image_folder  # Store the folder containing images
        self.text_file = text_file  # Store the Excel file containing text data (optional)
        self.sleep_time = tk.IntVar(value=sleep_time)
        self.image_files = None  # Initialize variable to store image files iterator
        self.text_data = None  # Initialize variable to store text data iterator
        self.current_image = None  # Initialize variable to store current image
        self.current_text = None  # Initialize variable to store current text
        self.image_label = tk.Label(master)  # Create a label widget for displaying images
        self.image_label.place(x=10, y=10)  # Place the image label in the window
        if text_file:
            self.text_label = tk.Label(master, wraplength=600, font=("Arial Bold", 20),justify=tk.LEFT)  # Create a label widget for displaying text
            self.text_label.place(x=730, y=10)  # Place the text label in the window
        self.switch_sheet()  # Start processing first sheet
        self.create_radio_buttons()

    def switch_sheet(self):
        # Method to switch between sheets
        excel_file = self.text_file
        if excel_file:
            # Open the Excel workbook
            pxl_doc = load_workbook(excel_file)
            sheet_names = pxl_doc.sheetnames
            self.image_files = self.load_images(pxl_doc, sheet_names)  # Load image files for all sheets
            self.text_data = self.load_text_data(pxl_doc, sheet_names)  # Load text data for all sheets
            self.load_next_slide()  # Load the first slide

    def load_images(self, pxl_doc, sheet_names):
        # Method to load image files from all sheets
        image_files = []
        for sheet_name in sheet_names:
            sheet = pxl_doc[sheet_name]
            image_loader = SheetImageLoader(sheet)
            for row in sheet.iter_rows(min_row=2, min_col=6, max_col=6):
                for cell in row:
                    image = image_loader.get(cell.coordinate)
                    if image:
                        image = image.convert('RGB')
                        image_files.append((f'image_{sheet_name}_{cell.coordinate}.jpg', image))
        return cycle(image_files)

    def load_text_data(self, pxl_doc, sheet_names):
        # Method to load text data from all sheets
        text_data = []
        for sheet_name in sheet_names:
            sheet = pxl_doc[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                formatted_row = []
                for header, value in zip(headers, row):
                    if value is not None:
                        formatted_row.append(f"{header} - {value}")
                text_data.append('\n\n'.join(header + '\n' + value for header, value in [pair.split(" - ") for pair in formatted_row]))
        return cycle(text_data)
    
    def create_radio_buttons(self):
        # Method to create radio buttons for selecting sleep time
        radio_frame = tk.Frame(self.master)  # Create a frame to hold the radio buttons
        radio_frame.place(x=900,y=650)  # Place the frame below the image label
        radio_label = tk.Label(radio_frame, text="Select Speed:",  font=("Aerial Bold" , 15))  # Create a label for the radio buttons
        radio_label.grid(column=0, row=0, padx=5, pady=5)  # Place the label in the frame
        radio_values = [1, 3, 5]  # Possible sleep time values
        for i, value in enumerate(radio_values):
            # Create a radio button for each sleep time value
            radio_button = tk.Radiobutton(radio_frame, text=f"{value}", variable=self.sleep_time, value=value ,)
            radio_button.grid(column=i+1, row=0, padx=5, pady=5)  # Place the radio button in the frame

    def load_next_slide(self):
        # Method to load the next slide
        try:
            self.current_image, image = next(self.image_files)  # Get the next image file
            if self.text_file:
                self.current_text = next(self.text_data)  # Get the next text (if available)
            image = image.resize((720, 600))  # Resize the image
            photo = ImageTk.PhotoImage(image)  # Create a PhotoImage object
            self.image_label.config(image=photo)  # Configure the image label with the new image
            self.image_label.image = photo  # Keep a reference to the PhotoImage object
            if self.text_file:
                self.text_label.config(text=self.current_text)  # Configure the text label with the new text (if available)
            self.master.title("Phoenix Safety Instructions")  # Set the title of the window
            self.master.after(self.sleep_time.get() * 1000, self.load_next_slide)  # Schedule loading of the next slide
        except StopIteration:
            # End of images or text data, reset to the beginning
               self.switch_sheet()  # Switch to the next sheet

    # Other methods remain the same

def browse_excel_file():
    # Function to browse for an Excel file
    filename = filedialog.askopenfilename(title="Select XL Sheet", filetypes=[("Excel Files", "*.xlsx")])  # Open a file dialog to select an Excel file
    if filename:
        return filename  # Return the selected filename
    else:
        return None  # Return None if no file is selected

def delete_images(image_folder):
    # Function to delete stored images
    image_files = [os.path.join(image_folder, f) for f in os.listdir(image_folder) if f.endswith(('.png', '.jpg', '.jpeg'))]  # Get a list of image files in the folder
    for file in image_files:  # Iterate through image files
        os.remove(file)  # Remove each image file
    print("Images deleted successfully.")  # Print a message indicating successful deletion

def start_slideshow():
    # Function to start the slideshow
    excel_file = browse_excel_file()  # Fetch Excel file
    if excel_file:
        image_folder = r"D:\Images"  # Change this to your image folder path

        # Fetch images from all sheets in Excel file
        pxl_doc = load_workbook(excel_file)
        
        for sheet_name in pxl_doc.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            sheet = pxl_doc[sheet_name]
            image_loader = SheetImageLoader(sheet)
            image_count = 0
            # Iterate through each row and fetch images
            for row in sheet.iter_rows(min_row=2, max_row=6, min_col=6, max_col=6):
                for cell in row:
                    image = image_loader.get(cell.coordinate) 
                    if image:
                        image = image.convert('RGB')       
                        image.save(f'D:/Images/image_{sheet_name}_{cell.coordinate}.jpg')
                        image_count += 1

            print(f"Images saved from sheet {sheet_name}: {image_count}")

        print("Images saved successfully.")

        # Pause execution for 5 seconds
        time.sleep(5)

        # Start slideshow
        root.withdraw()  # Hide the main window during the slideshow
        root_slideshow = tk.Toplevel(root)
        root_slideshow.geometry("1366x768")

        def on_close():
            delete_images(image_folder)
            root_slideshow.destroy()

        root_slideshow.protocol("WM_DELETE_WINDOW", on_close)

        app = SlideshowApp(root_slideshow, image_folder, excel_file)
        root_slideshow.mainloop()
        root.deiconify()  # Show the main window after the slideshow

def main():
    global root
    root = tk.Tk() 
    root.geometry("1366x768")  # Set initial window size
    root.title("Phoenix process Automation")
    img = tk.PhotoImage(file=r"C:\Users\USER\Pictures\Safety1.png")
    root.iconphoto(False, img)
    Style().configure("TFrame", background="#333")

   
    Label2 = tk.Label(root, text="Phoeneix Process Automation", font=(("Arial Bold", 25)))
    Label2.pack()
    Label1 = tk.Label(root, text="Safety Instructions", font=(("Arial Bold", 25)))
    Label1.pack()
    # Button to start slideshow
    start_button = tk.Button(root, text="Start",font=("Sans Bold" , 17) ,command=start_slideshow , fg="White" , bg="Blue")
    start_button.pack()
    image1 = Image.open(r"C:\Users\USER\Pictures\Backdrop.jpg")
    test = ImageTk.PhotoImage(image1)
    img1 = tk.Label(image=test)
    img1.image = test
    # Position image
    img1.pack( ipady=10)

    footer =  tk.Label(root ,text="Phoeneix Process Automation ©" , font=(("Arial Bold", 15)))
    footer.pack(side='bottom')

    root.mainloop()


if __name__ == "__main__":
    main()
